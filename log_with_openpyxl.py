from datetime import datetime
from openpyxl import workbook
from openpyxl.styles import NamedStyle
import time
from termcolor import colored
import openpyxl

# Conditional arguments required: users unique_id, transaction type, transaction amount.
# Logs the time of the transaction, users unique id, transcation type, transaction amount. saves in excel spreadsheet.
# Datetime is converted to string format so that it can be outputted in the correct format later in the program.
def logging_transaction(unique_id, type, amount):
    log_file_location = r"C:\Users\Vico Schot\Documents\Software Development Course\Group Banking App\Group Banking App\transaction_log.xlsx"
    log_file = openpyxl.load_workbook(log_file_location)
    current_sheet = log_file.active
    time_stamp = datetime.now()
    time_stamp_string_format = time_stamp.strftime("%m/%d/%Y at %H:%M:%S")
    current_sheet.cell(column=1, row=current_sheet.max_row+1, value=time_stamp_string_format)
    current_sheet.cell(column=2, row=current_sheet.max_row, value=unique_id)
    current_sheet.cell(column=3, row=current_sheet.max_row, value=type)
    current_sheet.cell(column=4, row=current_sheet.max_row, value=amount).number_format= '£0.00'
    log_file.save("transaction_log.xlsx")

# Outputs transaction history from user = unique_id
# Formats the output into a nice table, using termcolor to make headers bold, and transaction either green or red dependent on transaction type.
def output_user_specific_transactions(unique_id):
    log_file_location = r"C:\Users\Vico Schot\Documents\Software Development Course\Group Banking App\Group Banking App\transaction_log.xlsx"
    log_file = openpyxl.load_workbook(log_file_location)
    current_sheet = log_file.active
    print(colored("|          Time Stamp          |     Unique ID    |     Transaction Type    |    Amount    |", 'white', attrs=['bold']))
    for row in current_sheet.iter_rows(min_row=2, values_only=True):
        if row[1] == unique_id:
            if row[2] == 'Deposit':
                print(colored(f"|    {row[0]}    |        {row[1]}      |          {row[2]}        |     £{row[3]}     |", "green"))
            else:
                print(colored(f"|    {row[0]}    |        {row[1]}      |          {row[2]}       |     £{row[3]}     |", "red"))
